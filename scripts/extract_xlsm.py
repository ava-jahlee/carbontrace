"""
GHGCalc_V0m_lja.xlsm → JSON 덤프.

원본 엑셀의 모든 셀 (값·수식) 을 시트별로 그대로 뽑는다.
후속 스크립트가 이 raw JSON 을 읽어 TypeScript 데이터 모듈로 정제한다.

Usage:
    python scripts/extract_xlsm.py <path-to-xlsm> [out-dir]

Default:
    xlsm = C:/Workspace/My/온실가스/GHGCalc_V0m_lja.xlsm
    out  = src/data/raw/
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

DEFAULT_XLSM = Path("C:/Workspace/My/온실가스/GHGCalc_V0m_lja.xlsm")
DEFAULT_OUT = Path(__file__).resolve().parent.parent / "src" / "data" / "raw"


def dump_sheet(ws, with_formulas: bool):
    """시트 하나를 2차원 배열로 dump. 각 셀은 {v: value, f?: formula}."""
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    rows: list[list[object]] = []
    for r in range(1, max_row + 1):
        row: list[object] = []
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            v = cell.value
            entry: dict[str, object] = {}
            if with_formulas and isinstance(v, str) and v.startswith("="):
                entry["f"] = v
            else:
                entry["v"] = v
            row.append(entry if entry else None)
        rows.append(row)
    merged = [str(rng) for rng in ws.merged_cells.ranges]
    return {
        "max_row": max_row,
        "max_col": max_col,
        "max_col_letter": get_column_letter(max_col) if max_col else "",
        "merged": merged,
        "rows": rows,
    }


def main() -> int:
    xlsm_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSM
    out_dir = Path(sys.argv[2]) if len(sys.argv) > 2 else DEFAULT_OUT
    if not xlsm_path.exists():
        print(f"! xlsm not found: {xlsm_path}", file=sys.stderr)
        return 1
    out_dir.mkdir(parents=True, exist_ok=True)

    # 값과 수식을 각각 로드 (data_only=True 는 마지막 저장 시 계산된 값)
    wb_vals = openpyxl.load_workbook(xlsm_path, data_only=True, keep_vba=False)
    wb_form = openpyxl.load_workbook(xlsm_path, data_only=False, keep_vba=False)

    manifest = {
        "source": str(xlsm_path),
        "sheets": [],
    }

    for name in wb_vals.sheetnames:
        vals = dump_sheet(wb_vals[name], with_formulas=False)
        forms = dump_sheet(wb_form[name], with_formulas=True)
        combined_rows = []
        for r in range(vals["max_row"]):
            row_out = []
            for c in range(vals["max_col"]):
                v_entry = vals["rows"][r][c] if r < len(vals["rows"]) and c < len(vals["rows"][r]) else None
                f_entry = forms["rows"][r][c] if r < len(forms["rows"]) and c < len(forms["rows"][r]) else None
                cell = {}
                if v_entry and v_entry.get("v") is not None:
                    cell["v"] = v_entry["v"]
                if f_entry and f_entry.get("f") is not None:
                    cell["f"] = f_entry["f"]
                combined_rows.append(cell if cell else None)
            # keep as flat list per row length? no, keep as 2d
        # redo cleanly (2d)
        rows_2d = []
        for r in range(vals["max_row"]):
            row_2d = []
            for c in range(vals["max_col"]):
                v_entry = vals["rows"][r][c] if r < len(vals["rows"]) and c < len(vals["rows"][r]) else None
                f_entry = forms["rows"][r][c] if r < len(forms["rows"]) and c < len(forms["rows"][r]) else None
                cell: dict[str, object] = {}
                if v_entry and v_entry.get("v") is not None:
                    cell["v"] = v_entry["v"]
                if f_entry and f_entry.get("f") is not None:
                    cell["f"] = f_entry["f"]
                row_2d.append(cell if cell else None)
            rows_2d.append(row_2d)

        sheet_dump = {
            "name": name,
            "max_row": vals["max_row"],
            "max_col": vals["max_col"],
            "max_col_letter": vals["max_col_letter"],
            "merged": vals["merged"],
            "rows": rows_2d,
        }
        safe_name = name.replace("/", "_").replace("\\", "_")
        out_file = out_dir / f"sheet_{safe_name}.json"
        with out_file.open("w", encoding="utf-8") as fp:
            json.dump(sheet_dump, fp, ensure_ascii=False, indent=2)
        manifest["sheets"].append({
            "name": name,
            "file": out_file.name,
            "rows": sheet_dump["max_row"],
            "cols": sheet_dump["max_col"],
            "cols_letter": sheet_dump["max_col_letter"],
        })
        print(f"[ok] {name:20s}  {sheet_dump['max_row']:>4} x {sheet_dump['max_col']:>3}  -> {out_file.name}")

    with (out_dir / "manifest.json").open("w", encoding="utf-8") as fp:
        json.dump(manifest, fp, ensure_ascii=False, indent=2)

    print(f"\nwrote {len(manifest['sheets'])} sheets to {out_dir}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

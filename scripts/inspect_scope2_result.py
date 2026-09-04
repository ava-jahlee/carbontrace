"""xlsm Main 시트의 Scope 2 계산 결과 값 조회 (파리티용)."""
from openpyxl import load_workbook
from pathlib import Path

XLSM = Path(r"C:\Workspace\Private\온실가스\GHGCalc_V0m_lja.xlsm")

wb = load_workbook(XLSM, data_only=True)
ws = wb["Main"]

# Main 시트 Scope 2 부분 값 (data_only=True 라 캐시된 계산값)
print("=" * 60)
print("Main Scope 2 셀 값 (data_only)")
print("=" * 60)
targets = [
    ("C31", "연료 선택"),
    ("D31", "GHG 선택"),
    ("E31", "사용량"),
    ("F31", "단위"),
    ("G33", "지역"),
    ("H33", "계획기간"),
    ("D39", "전력 배출계수 (kg/MWh)"),
    ("E39", "열 배출계수 (kg/TJ)"),
    ("F39", "GWP"),
    ("D42", "tGHG"),
    ("E42", "tCO2eq"),
    ("Main!F7", "전력 배출계수 판"),
    ("Main!G7", "전력 순발열량 판"),
]
for coord, label in targets:
    c = coord.replace("Main!", "")
    v = ws[c].value
    print(f"  [{coord}] {label}: {v}")

"""_Law&GL22 시트의 전력 배출계수 부분 (B100:U103) 조사."""
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pathlib import Path

XLSM = Path(r"C:\Workspace\Private\온실가스\GHGCalc_V0m_lja.xlsm")
OUT = Path("docs/refs/scope2-electricity.txt")

wb_v = load_workbook(XLSM, data_only=True)
wb_f = load_workbook(XLSM, data_only=False)

lines: list[str] = []
def w(s: str = ""):
    lines.append(s)

for sheet_name in ["_Law&GL22"]:
    w("=" * 80)
    w(f"Sheet: {sheet_name}, 전력 배출계수 부분 (B95:U110)")
    w("=" * 80)
    ws_v = wb_v[sheet_name]
    ws_f = wb_f[sheet_name]

    # 헤더 (95~99행)
    for r in range(95, 111):
        for c in range(1, 25):
            vv = ws_v.cell(row=r, column=c).value
            vf = ws_f.cell(row=r, column=c).value
            if vv is None and vf is None:
                continue
            coord = f"{get_column_letter(c)}{r}"
            vf_s = str(vf) if vf is not None else ""
            vv_s = str(vv) if vv is not None else ""
            if vf_s == vv_s:
                w(f"  [{coord}] {vv_s[:100]}")
            else:
                w(f"  [{coord}] formula={vf_s[:80]} | value={vv_s[:80]}")
        w("")  # 행 구분

OUT.write_text("\n".join(lines), encoding="utf-8")
print(f"Wrote {len(lines)} lines to {OUT}")

"""
GHGCalc_V0m_lja.xlsm · 시설 (사업장) 진입 스키마 추출.

이 스크립트가 뽑아내는 것:
1. Main 시트의 시설 관련 셀 (E2/F2/G2/H2 · 유형·분야·부문·용도)
   - 실제 값 · 함수식
2. Main!E3~H5 등 사용자 입력 위치의 수식 · vlookup 참조 범위
3. Main!D14/E14/F14 등 · 등급(F5) → 최소 Tier 매핑 수식
4. _Law&GL22!B109:C112 등급 기준 (연간 GHG 만ton/yr 범위)
5. xlsm 내부의 Data Validation XML 직접 파싱 (openpyxl 미지원)
   - dropdown 후보 값을 xl/worksheets/sheet1.xml 에서 추출

출력 · docs/refs/facility_schema.txt (UTF-8)
"""

from __future__ import annotations

import sys
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

sys.stdout.reconfigure(encoding="utf-8")

from openpyxl import load_workbook

XLSM = Path(r"C:\Workspace\Private\온실가스\GHGCalc_V0m_lja.xlsm")
OUT = Path(r"C:\Workspace\active\carbontrace\docs\refs\facility_schema.txt")
OUT.parent.mkdir(parents=True, exist_ok=True)

lines: list[str] = []


def w(s: str = "") -> None:
    lines.append(s)
    print(s)


w("=" * 72)
w("Facility (시설 · 사업장) 진입 스키마 추출")
w("=" * 72)

# ─── openpyxl 로 셀 값·수식 확인 ────────────────────────────────
wb = load_workbook(XLSM, data_only=False, keep_vba=True)
wb_val = load_workbook(XLSM, data_only=True, keep_vba=True)

main = wb["Main"]
main_val = wb_val["Main"]
law = wb["_Law&GL22"]
law_val = wb_val["_Law&GL22"]


def dump_cell(sheet, sheet_val, coord: str, label: str = "") -> None:
    v = sheet[coord].value
    vv = sheet_val[coord].value
    prefix = f"{coord}"
    if label:
        prefix += f" [{label}]"
    w(f"  {prefix:22s} · formula={v!r}")
    if v != vv:
        w(f"  {'':22s} · value  ={vv!r}")


w("")
w("── Main 시트 · 시설 정보 영역 (row 2 ~ 5) ───────────────────")

for r in range(2, 6):
    w("")
    w(f"  · row {r}:")
    for c_letter in "ABCDEFGHIJKLMNOPQ":
        coord = f"{c_letter}{r}"
        v = main[coord].value
        vv = main_val[coord].value
        if v is None:
            continue
        pref = f"    {coord}"
        if isinstance(v, str) and v.startswith("="):
            w(f"    {coord:6s} formula = {v!r}")
            w(f"    {'':6s} value   = {vv!r}")
        else:
            w(f"    {coord:6s} {v!r}")


w("")
w("── Main!F5 (등급 자동 산정) · D14/E14/F14/D15/E15/F15 (Tier) ─")

for coord, label in [
    ("F5", "등급 A/B/C 자동"),
    ("G5", "연간 GHG · 만ton/yr · 사용자 입력"),
    ("D14", "열량계수 최소 Tier"),
    ("E14", "배출계수 최소 Tier"),
    ("F14", "산화계수 최소 Tier"),
    ("D15", "열량계수 사용 Tier"),
    ("E15", "배출계수 사용 Tier"),
    ("F15", "산화계수 사용 Tier"),
]:
    dump_cell(main, main_val, coord, label)


w("")
w("── _Law&GL22 시트 · 등급 기준 (row 108 ~ 112) ────────────────")

for r in range(107, 113):
    w("")
    w(f"  · row {r}:")
    for c_letter in "ABCDEFGH":
        coord = f"{c_letter}{r}"
        v = law[coord].value
        vv = law_val[coord].value
        if v is None:
            continue
        w(f"    {coord:6s} {v!r}  (val={vv!r})")


# ─── Data Validation · xlsm 내부 XML 직접 파싱 ──────────────────

w("")
w("=" * 72)
w("xlsm 내부 XML · Data Validation dropdown 목록")
w("=" * 72)

ns = {
    "s": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "x14": "http://schemas.microsoft.com/office/spreadsheetml/2009/9/main",
    "xm": "http://schemas.microsoft.com/office/excel/2006/main",
}

with zipfile.ZipFile(XLSM, "r") as z:
    # 시트 이름 → 파일 매핑
    wb_xml = z.read("xl/workbook.xml").decode("utf-8")
    wb_root = ET.fromstring(wb_xml)
    sheet_files: dict[str, str] = {}
    rels = z.read("xl/_rels/workbook.xml.rels").decode("utf-8")
    rels_root = ET.fromstring(rels)
    rel_map = {
        r.attrib["Id"]: r.attrib["Target"]
        for r in rels_root.findall("{http://schemas.openxmlformats.org/package/2006/relationships}Relationship")
    }
    for sh in wb_root.findall("s:sheets/s:sheet", ns):
        name = sh.attrib["name"]
        rid = sh.attrib["{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"]
        target = rel_map[rid]
        if not target.startswith("xl/"):
            target = "xl/" + target
        sheet_files[name] = target

    w("")
    w("  sheet 파일 매핑:")
    for k, v in sheet_files.items():
        w(f"    {k} · {v}")

    # Main 시트 XML 파싱
    main_path = sheet_files["Main"]
    main_xml = z.read(main_path).decode("utf-8")
    main_root = ET.fromstring(main_xml)

    w("")
    w(f"── Main 시트 dataValidations · 표준 위치 ────────────────────")

    dvs = main_root.find("s:dataValidations", ns)
    if dvs is not None:
        for dv in dvs.findall("s:dataValidation", ns):
            typ = dv.attrib.get("type", "?")
            sqref = dv.attrib.get("sqref", "?")
            formula = dv.find("s:formula1", ns)
            f_text = formula.text if formula is not None else "?"
            w(f"    type={typ:12s} sqref={sqref:20s} formula1={f_text!r}")

    # x14 확장 dataValidations (더 긴 목록)
    w("")
    w(f"── Main 시트 dataValidations · x14 확장 ────────────────────")

    ext_lst = main_root.find("s:extLst", ns)
    if ext_lst is not None:
        for ext in ext_lst.findall("s:ext", ns):
            x14_dvs = ext.find("x14:dataValidations", ns)
            if x14_dvs is not None:
                for dv in x14_dvs.findall("x14:dataValidation", ns):
                    typ = dv.attrib.get("type", "?")
                    formula = dv.find("x14:formula1", ns)
                    if formula is not None:
                        f_ref = formula.find("xm:f", ns)
                        f_text = f_ref.text if f_ref is not None else "?"
                    else:
                        f_text = "?"
                    sqref_elem = dv.find("xm:sqref", ns)
                    sqref = sqref_elem.text if sqref_elem is not None else "?"
                    w(f"    type={typ:12s} sqref={sqref:20s} formula1={f_text!r}")

    # _Law&GL22 시트 · dropdown 값이 실제로 여기에 있을 것 (참조)
    w("")
    w("── _Law&GL22 시트 · 상단 (row 1~34) · dropdown 후보 원본 ───")

    law_path = sheet_files["_Law&GL22"]
    # openpyxl 로 다시 · 셀 값 dump (row 1~34 · col A~V)
    for r in range(1, 35):
        row_cells = []
        for c in range(1, 23):
            v = law.cell(row=r, column=c).value
            if v is not None and (not isinstance(v, str) or v.strip() != ""):
                coord = law.cell(row=r, column=c).coordinate
                row_cells.append(f"{coord}={v!r}")
        if row_cells:
            w(f"  row {r:2d}: " + " | ".join(row_cells[:8]))

w("")
w("=" * 72)
w("리포트 종료")
w("=" * 72)

OUT.write_text("\n".join(lines), encoding="utf-8")
print(f"\n[✓] 저장: {OUT}")

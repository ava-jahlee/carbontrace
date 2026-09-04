"""
GHGCalc_V0m_lja.xlsm 을 감사자 관점이 아닌 · 사용자 흐름 관점으로 재조사.

목적:
- 실제 사용자가 이 엑셀을 어떻게 쓰는지 (인벤토리? 합산? 저장?)
- 시트 목록 + 각 시트의 성격 · 데이터 위치 · 함수 구조
- 입력 시트 · 합산 시트 · 참조 시트 구분
- 총합·소계 방식 파악

출력:
- docs/refs/xlsm_flow_audit.txt (UTF-8) · 사용자 흐름 관점 리포트
"""

from __future__ import annotations

import os
import sys
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")

from openpyxl import load_workbook

XLSM = Path(r"C:\Workspace\Private\온실가스\GHGCalc_V0m_lja.xlsm")
OUT = Path(r"C:\Workspace\active\carbontrace\docs\refs\xlsm_flow_audit.txt")

if not XLSM.exists():
    print(f"[!] xlsm 파일을 찾지 못했음: {XLSM}")
    sys.exit(1)

OUT.parent.mkdir(parents=True, exist_ok=True)
lines: list[str] = []

def w(s: str = "") -> None:
    lines.append(s)
    print(s)

w("=" * 72)
w(f"GHGCalc_V0m_lja.xlsm · 사용자 흐름 관점 재감사")
w("=" * 72)

# 함수 문자열도 필요 · data_only=False (openpyxl 기본)
wb = load_workbook(XLSM, data_only=False, keep_vba=True)
# 계산된 값도 병행 · data_only=True
wb_val = load_workbook(XLSM, data_only=True, keep_vba=True)

w("")
w(f"시트 총 {len(wb.sheetnames)} 개")
w("")

for i, name in enumerate(wb.sheetnames):
    ws = wb[name]
    max_r = ws.max_row
    max_c = ws.max_column
    hidden = ws.sheet_state
    tab_color = ws.sheet_properties.tabColor
    w(f"  {i+1:2d}. [{name}]  · state={hidden}  · rows={max_r}  · cols={max_c}"
      + (f"  · tab={tab_color.rgb if tab_color else '-'}" if tab_color else ""))

w("")
w("=" * 72)
w("각 시트 · 상위 셀 스캔 (첫 20 rows · 8 cols · header · 소계·합계 후보)")
w("=" * 72)

def is_meaningful(v):
    if v is None:
        return False
    if isinstance(v, str) and v.strip() == "":
        return False
    return True

for name in wb.sheetnames:
    ws = wb[name]
    ws_val = wb_val[name]
    max_r = min(ws.max_row, 40)
    max_c = min(ws.max_column, 12)

    w("")
    w(f"─── [{name}] " + "─" * (68 - len(name)))
    w(f"  state={ws.sheet_state} · dims=A1:{ws.cell(row=ws.max_row, column=ws.max_column).coordinate}")

    # 상위 rows 스캔 · header 후보
    for r in range(1, max_r + 1):
        row_cells = []
        for c in range(1, max_c + 1):
            cell = ws.cell(row=r, column=c)
            cell_v = ws_val.cell(row=r, column=c)
            v = cell.value
            if is_meaningful(v):
                # 함수라면 f= 로 표기 · 값이 있으면 = 값
                if isinstance(v, str) and v.startswith("="):
                    row_cells.append(f"{cell.coordinate}=f{v}")
                else:
                    row_cells.append(f"{cell.coordinate}={v!r}")
        if row_cells:
            w("  " + " | ".join(row_cells[:6]))  # 첫 6 셀만

    # 소계·합계 후보 · SUM · 소계 · 합계 · 총합 검색
    subtotal_hits = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            s = str(cell.value)
            if any(k in s for k in ("합계", "소계", "총합", "SUM(", "SUBTOTAL(", "Total", "TOTAL")):
                subtotal_hits.append((cell.coordinate, s[:100]))
                if len(subtotal_hits) >= 8:
                    break
        if len(subtotal_hits) >= 8:
            break

    if subtotal_hits:
        w(f"  · 소계/합계 후보 ({len(subtotal_hits)}건 상위):")
        for coord, s in subtotal_hits:
            w(f"    {coord}: {s}")

w("")
w("=" * 72)
w("리포트 종료")
w("=" * 72)

OUT.write_text("\n".join(lines), encoding="utf-8")
print(f"\n[✓] 저장: {OUT}")

"""build_scope2_data.py — Scope 2 (전력·열) 배출계수 데이터 생성.

원본: C:\\Workspace\\Private\\온실가스\\GHGCalc_V0m_lja.xlsm
  - _Law&GL22!J99~L100 : 전력 배출계수 (2017년/2022년 승인 판)
  - _Law&GL22!J101~L103 : 열(스팀) 3종 (열전용/열병합/열평균, 원출처 미상)
  - _Supplier!C6~E21 : KDHC 지사별 열 배출계수 (16개 = 8지사 × 계획기간 3기·4기)

출력: src/data/factors/scope2.gen.ts

사용법:
    python -X utf8 scripts/build_scope2_data.py
"""
from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import load_workbook

# ── 경로 ─────────────────────────────────────────────────────
XLSM = Path(r"C:\Workspace\Private\온실가스\GHGCalc_V0m_lja.xlsm")
OUT_TS = Path("src/data/factors/scope2.gen.ts")

# ── xlsm 원본 값 (하드코딩 · 자동추출 실패 대비 인라인 문서화) ────

# 전력 배출계수 (kg/MWh, MWh 단위 유의)
POWER_EF = {
    "2017": {
        "발전단": {"CO2": 440.1, "CH4": 0.0034, "N2O": 0.0082},
        "소비단": {"CO2": 456.7, "CH4": 0.0036, "N2O": 0.0085},
    },
    "2022": {
        "발전단": {"CO2": 440.3, "CH4": 0.0116, "N2O": 0.0093},
        "소비단": {"CO2": 474.7, "CH4": 0.0125, "N2O": 0.01},
    },
}
POWER_NCV = {"발전단": 8.9, "소비단": 9.6}  # MJ/kWh (별표 12 표 A 와 동일)

# 열(스팀) 국가 통합 3종 (kgGHG/TJ, 원출처 미상)
HEAT_NATIONAL = {
    "열전용": {"CO2": 56373, "CH4": 1.278, "N2O": 0.166},
    "열병합": {"CO2": 60760, "CH4": 2.053, "N2O": 0.549},
    "열평균": {"CO2": 59510, "CH4": 1.832, "N2O": 0.44},
}

# KDHC 지사별 배출계수 (kgGHG/TJ) — _Supplier 시트에서 자동 추출
DISTRICTS = [
    "수도권지사", "평택지사", "청주지사", "세종지사",
    "대구지사", "양산지사", "김해지사", "광주전남지사",
]


def extract_kdhc_from_xlsm() -> dict:
    """xlsm _Supplier 시트에서 지사별·계획기간별 값 자동 추출."""
    wb = load_workbook(XLSM, data_only=True)
    ws = wb["_Supplier"]

    result: dict[str, dict[str, dict[str, float]]] = {"3기": {}, "4기": {}}

    # 3기: 6~13행, 4기: 14~21행
    for r_start, phase in [(6, "3기"), (14, "4기")]:
        for idx, district in enumerate(DISTRICTS):
            r = r_start + idx
            district_cell = ws.cell(row=r, column=2).value  # B열
            assert district_cell == district, f"row {r}: expected {district}, got {district_cell}"
            result[phase][district] = {
                "CO2": float(ws.cell(row=r, column=3).value),  # C
                "CH4": float(ws.cell(row=r, column=4).value),  # D
                "N2O": float(ws.cell(row=r, column=5).value),  # E
            }
    return result


# ── TS 렌더링 유틸 ───────────────────────────────────────────

TS_HEADER = """/**
 * scope2.gen.ts — Scope 2 배출계수 (자동 생성 · 손대지 마세요)
 *
 * 원본: C:\\\\Workspace\\\\Private\\\\온실가스\\\\GHGCalc_V0m_lja.xlsm
 * 생성: scripts/build_scope2_data.py
 *
 * 카테고리:
 *   POWER     : 전력 배출계수 (kg/MWh, 발전단/소비단, GIR 승인 연도별)
 *   HEAT_KDHC : 한국지역난방공사 지사별 열 배출계수 (kg/TJ, 8지사 × 계획기간 3기/4기)
 *   HEAT_NATIONAL : 국가 통합 열(스팀) 3종 (kg/TJ, 열전용/열병합/열평균) - 원출처 미상
 */

import type { Measurement } from "@/data/factors/types";
import {
  GIR_POWER_2017,
  GIR_POWER_2022,
  KDHC_HEAT_EF,
  KETS_HEAT_EF,
} from "@/data/sources";

export type PowerVintage = "2017" | "2022";
export type PowerLocation = "발전단" | "소비단";
export type PowerGhg = "CO2" | "CH4" | "N2O";
export type HeatKind = "열전용" | "열병합" | "열평균";
export type KdhcPhase = "3기" | "4기";
export type KdhcDistrict =
  | "수도권지사" | "평택지사" | "청주지사" | "세종지사"
  | "대구지사" | "양산지사" | "김해지사" | "광주전남지사";

export interface PowerEmissionFactors {
  vintage: PowerVintage;
  location: PowerLocation;
  ncv: Measurement;                       // MJ/kWh
  ef: Record<PowerGhg, Measurement>;      // kg/MWh
}

export interface HeatEmissionFactorsKdhc {
  phase: KdhcPhase;
  district: KdhcDistrict;
  ef: Record<PowerGhg, Measurement>;      // kg/TJ
}

export interface HeatEmissionFactorsNational {
  kind: HeatKind;
  ef: Record<PowerGhg, Measurement>;      // kg/TJ
}

"""


def render_measurement(value, unit: str, source_ref: str, note: str | None = None) -> str:
    """Measurement TS 리터럴 렌더링."""
    ps = f"{{ ...{source_ref}"
    if note:
        ps += f', "note": {json.dumps(note, ensure_ascii=False)}'
    ps += " }"
    return (
        f'{{ "value": {json.dumps(value)}, "unit": {json.dumps(unit)}, '
        f'"primarySource": {ps} }}'
    )


def build_power_ts() -> str:
    lines = ["export const POWER: PowerEmissionFactors[] = ["]
    for vintage in ["2017", "2022"]:
        source_ref = f"GIR_POWER_{vintage}"
        for location in ["발전단", "소비단"]:
            ncv_val = POWER_NCV[location]
            ncv_m = render_measurement(
                ncv_val, "MJ/kWh", source_ref,
                note=f"{vintage}년 승인 · {location} 순발열량 (K-ETS 별표 12 표 A 전기 항목과 동일)",
            )
            ef_lines = []
            for ghg in ["CO2", "CH4", "N2O"]:
                v = POWER_EF[vintage][location][ghg]
                m = render_measurement(
                    v, "kgGHG/MWh", source_ref,
                    note=f"{vintage}년 승인 · {location} · {ghg}",
                )
                ef_lines.append(f'    "{ghg}": {m},')
            lines.append("  {")
            lines.append(f'    "vintage": "{vintage}",')
            lines.append(f'    "location": "{location}",')
            lines.append(f'    "ncv": {ncv_m},')
            lines.append('    "ef": {')
            for el in ef_lines:
                lines.append("  " + el)
            lines.append("    },")
            lines.append("  },")
    lines.append("];")
    return "\n".join(lines)


def build_heat_kdhc_ts(kdhc_data: dict) -> str:
    lines = ["export const HEAT_KDHC: HeatEmissionFactorsKdhc[] = ["]
    for phase in ["3기", "4기"]:
        for district in DISTRICTS:
            vals = kdhc_data[phase][district]
            ef_lines = []
            for ghg in ["CO2", "CH4", "N2O"]:
                m = render_measurement(
                    vals[ghg], "kgGHG/TJ", "KDHC_HEAT_EF",
                    note=f"KDHC 계획기간 {phase} · {district} · {ghg} · 실측 기반",
                )
                ef_lines.append(f'    "{ghg}": {m},')
            lines.append("  {")
            lines.append(f'    "phase": "{phase}",')
            lines.append(f'    "district": "{district}",')
            lines.append('    "ef": {')
            for el in ef_lines:
                lines.append("  " + el)
            lines.append("    },")
            lines.append("  },")
    lines.append("];")
    return "\n".join(lines)


def build_heat_national_ts() -> str:
    lines = ["export const HEAT_NATIONAL: HeatEmissionFactorsNational[] = ["]
    for kind in ["열전용", "열병합", "열평균"]:
        vals = HEAT_NATIONAL[kind]
        ef_lines = []
        for ghg in ["CO2", "CH4", "N2O"]:
            m = render_measurement(
                vals[ghg], "kgGHG/TJ", "KETS_HEAT_EF",
                note=f"⚠ 원출처 미상 · {kind} · {ghg} · xlsm 하드코딩값 (K-ETS 별표 또는 GIR 부속 자료 추정)",
            )
            ef_lines.append(f'    "{ghg}": {m},')
        lines.append("  {")
        lines.append(f'    "kind": "{kind}",')
        lines.append('    "ef": {')
        for el in ef_lines:
            lines.append("  " + el)
        lines.append("    },")
        lines.append("  },")
    lines.append("];")
    return "\n".join(lines)


def main():
    print("[extract] KDHC 지사별 배출계수 추출 중...")
    kdhc_data = extract_kdhc_from_xlsm()
    print(f"[extract] 3기 {len(kdhc_data['3기'])}개 지사, 4기 {len(kdhc_data['4기'])}개 지사")

    ts = (
        TS_HEADER
        + build_power_ts() + "\n\n"
        + build_heat_kdhc_ts(kdhc_data) + "\n\n"
        + build_heat_national_ts() + "\n"
    )

    OUT_TS.parent.mkdir(parents=True, exist_ok=True)
    OUT_TS.write_text(ts, encoding="utf-8")
    print(f"[write] {OUT_TS}")
    print(f"        POWER: {len(POWER_EF) * 2} entries (전력 {len(POWER_EF)} 판 × 발전단/소비단)")
    print(f"        HEAT_KDHC: {sum(len(v) for v in kdhc_data.values())} entries (계획기간 2판 × 지사 8개)")
    print(f"        HEAT_NATIONAL: {len(HEAT_NATIONAL)} entries (열전용/열병합/열평균)")


if __name__ == "__main__":
    main()
